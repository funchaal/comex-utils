import json
import requests
import zipfile
import io

from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO

import datetime
import os

import tkinter as tk
from tkinter import filedialog

import pandas as pd

import re

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side

def makeFillSheet(products, relation_json):

    data = []
    attr_values = []

    # def normalize_column_name(name):
    #     return re.sub(r'\s+', '_', name.strip().lower())

    # products.columns = [normalize_column_name(col) for col in products.columns]

    for _, row in products.iterrows():
        ncm = str(row['NCM'])
        raiz = row['Raíz']
        descricao = row['Descrição']
        denominacao = row['Denominação']
        codigo = row['Código Interno']
        modalidade = row['Modalidade']
        situacao = row['Situação']

        for i in range(8 - len(ncm)):
            ncm = '0' + ncm

        base_info = {
            'NCM': ncm,
            'Raíz (CNPJ)': raiz,
            'Código Interno': codigo,
            'Descrição': descricao,
            'Denominação': denominacao,
            'Modalidade': modalidade,
            'Situação': situacao, 
            'Atributos': []
        }

        attributes = next((a['listaAtributos'] for a in relation_json['listaNcm'] if str(a['codigoNcm']).replace('.', '') == str(ncm)), None)
        attributes_codes = [n['codigo'] for n in attributes]

        for attr in attributes:
            attr_dtls = next((d for d in relation_json['detalhesAtributos'] if d['codigo'] == attr['codigo']), None)

            attr_code = attr_dtls['codigo']
            attr_name = attr_dtls['nomeApresentacao']

            column_name = f'{attr_code} - {attr_name}'

            orientacaoPreenchimento = attr_dtls.get('orientacaoPreenchimento', '')

            if orientacaoPreenchimento:
                column_name = column_name + '\n\n' + orientacaoPreenchimento

            formaPreenchimento = attr_dtls.get('formaPreenchimento', '')

            if formaPreenchimento:
                column_name = column_name + '\n\nTipo preenchimento: ' + str(formaPreenchimento).replace('BOOLEANO', 'SIM ou NÃO')

            if formaPreenchimento == 'LISTA_ESTATICA':
                options = '\n\nPreencha um dos códigos abaixo:\n\n'
                for option in attr_dtls['dominio']:
                    options = options + option['codigo'] + ' - ' + option['descricao'] + '\n'
                
                column_name = column_name + options

            if attr['obrigatorio']:
                field_type = 1
            else:
                field_type = 3

            value = None

            name = None

            for col in products.columns:
                parts = col.split('-')
                if not (len(parts) > 0 and 'ATT' in parts[0]):
                    parts = col.split('\n')
                    if len(parts) > 0:
                        name = parts[0].strip().lower()
                    else:
                        name = col

                if attr_code.lower() in parts[0].lower() or (name and attr_name.lower() == name):
                    value = row[col]
            
            base_info['Atributos'].append({
                'column name': column_name, 
                'field type': field_type, 
                'value': value
            })

            if attr_dtls['atributoCondicionante']:
                for attr_cond in attr_dtls['condicionados']:
                    attr_cond_code = attr_cond['atributo']['codigo']
                    attr_cond_name = attr_cond['atributo']['nomeApresentacao']

                    cond_column_name = f'{attr_cond_code} - {attr_cond_name}'

                    orientacaoPreenchimento = attr_cond['atributo'].get('orientacaoPreenchimento', '')

                    if orientacaoPreenchimento:
                        cond_column_name = cond_column_name + ' - ' + orientacaoPreenchimento

                    formaPreenchimento = attr_cond['atributo'].get('formaPreenchimento', '')

                    if formaPreenchimento:
                        cond_column_name = cond_column_name + '\n\nTipo preenchimento: ' + str(formaPreenchimento).replace('BOOLEANO', 'SIM ou NÃO')

                    if formaPreenchimento == 'LISTA_ESTATICA':
                        options = '\n\nPreencha um dos códigos abaixo:\n\n'
                        for option in attr_cond['atributo']['dominio']:
                            options = options + option['codigo'] + ' - ' + option['descricao'] + '\n'
                        
                        cond_column_name = cond_column_name + options

                    cond_column_name = cond_column_name + f'\n\n(atributo condicional, preencha apenas se {attr_cond['descricaoCondicao']})'
                    cond_field_type = 2
                    cond_value = None

                    name = None

                    for col in products.columns:
                        parts = col.split('-')
                        if not (len(parts) > 0 and 'ATT' in parts[0]):
                            parts = col.split('\n')
                            if len(parts) > 0:
                                name = parts[0].strip().lower()
                            else:
                                name = col
                            
                        if attr_cond_code.lower() in col.lower() or attr_cond_name.lower() == name:
                            cond_value = row[col]
                    
                    base_info['Atributos'].append({
                        'column name': cond_column_name, 
                        'field type': cond_field_type, 
                        'value': cond_value
                    })

        data.append(base_info)

    wb = convert_to_excel(data)

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    return file_stream

def convert_to_excel(data):
    # Identificar todas as colunas fixas e dinâmicas
    fixed_columns = ["NCM", "Raíz (CNPJ)", "Descrição", "Denominação", "Código Interno", "Modalidade", "Situação"]
    attribute_columns = []
    
    for row in data:
        for attr in row.get("Atributos", []):
            if attr["column name"] not in attribute_columns:
                attribute_columns.append(attr["column name"])
    
    all_columns = fixed_columns + attribute_columns
    
    # Criar DataFrame vazio
    df = pd.DataFrame(columns=all_columns)
    
    # Popular DataFrame
    records = []
    attribute_styles = {}
    
    for row in data:
        record = {col: row.get(col, "") for col in fixed_columns}
        row_attributes = {attr["column name"]: attr for attr in row.get("Atributos", [])}
        
        for col_name in attribute_columns:
            if col_name in row_attributes:
                attr = row_attributes[col_name]
                record[col_name] = attr["value"]
                attribute_styles[(len(records) + 1, all_columns.index(col_name) + 1)] = attr["field type"]
            else:
                record[col_name] = ""
                attribute_styles[(len(records) + 1, all_columns.index(col_name) + 1)] = 0  # Vermelho para ausência do atributo
        
        records.append(record)
    
    df = pd.DataFrame(records, columns=all_columns)
    
    # Criar um novo Workbook e preencher com os dados do DataFrame
    wb = Workbook()
    ws = wb.active
    ws.append(all_columns)
    
    for row in records:
        ws.append([row.get(col, "") for col in all_columns])
    
    color_mapping = {
        1: "FFFFFF",  # Branco
        2: "FFFF00",  # Amarelo
        3: "D3D3D3",  # Cinza Claro
        0: "FF0000"   # Vermelho para ausência do atributo
    }

    border_style = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    for (row, col), field_type in attribute_styles.items():
        cell = ws.cell(row=row+1, column=col)
        cell.fill = PatternFill(start_color=color_mapping.get(field_type, "FFFFFF"), fill_type="solid")
        cell.border = border_style
    
    return wb


def get_relation_json(prod=False, set_token=None, csrf_token=None):
    url = None

    if prod:
        url = 'https://portalunico.siscomex.gov.br/cadatributos/api/atributo-ncm/download/json'
    else:
        url = 'https://val.portalunico.siscomex.gov.br/cadatributos/api/atributo-ncm/download/json'

    response = requests.get(url)
    if response.status_code == 200:
        with zipfile.ZipFile(io.BytesIO(response.content)) as z:
            # Obtém o único arquivo dentro do ZIP
            nome_arquivo = z.namelist()[0]  
            
            # Abre o JSON dentro do ZIP
            with z.open(nome_arquivo) as json_file:
                data = json.load(json_file)  # Lê o JSON
            
            return data
        return None
    else:
        response.raise_for_status()

def excel_to_dict(df):
    result = []
    for _, row in df.iterrows():
        ncm = str(row['ncm'])
        descricao = str(row['descrição'])
        denominacao = str(row['denominação'])
        raiz = str(row['raíz'])
        situacao = str(row['situação'])
        modalidade = str(row['modalidade'])
        codigointerno = str(row['código interno'])
        # codigo = str(row['codigo']).replace('.0', '')
        atributos = []
        for col in df.columns:
            if col not in ('ncm', 'descrição', 'denominação', 'código interno', 'raíz', 'situação', 'modalidade'):
                match = re.match(r'^(ATT_\d+)-(.*)$', col)
                if match:
                    code = match.group(1).strip()
                    name = match.group(2).strip().lower()
                else:
                    code = None
                    name = col
                atributos.append({'code': str(code).strip(), 'name': str(name).strip().lower(), 'value': str(row[col]).strip()})

        result.append({
            'descricao': descricao, 
            'denominacao': denominacao, 
            # 'codigo': codigo, 
            'codigoInterno': codigointerno, 
            'raiz': raiz, 
            'situacao': situacao, 
            'modalidade': modalidade, 
            'ncm': ncm, 
            'atributos': atributos
        })
    
    return result

def makeProductsRetify(products, attributes_json):
    
    results = []

    errors = []

    seq = 0

    for product in products:

        seq += 1

        ncm = str(product['ncm']).replace('.', '')
        descricao = str(product['descricao'])
        denominacao = str(product['denominacao'])
        raiz = (str(product['raiz']).replace('.', ''))[:8]
        situacao = str(product['situacao']).upper()
        modalidade = str(product['modalidade']).upper()
        codigointerno = str(product['codigoInterno']).strip()
        codigo = str(product['codigo']).strip()

        attribute_array = product['atributos']

        results.append({
            'seq': seq, 
            'descricao': descricao, 
            'codigo': codigo, 
            'denominacao': denominacao, 
            'cpfCnpjRaiz': raiz, 
            'situacao': 'DESATIVADO', 
            'modalidade': 'IMPORTACAO', 
            'ncm': ncm, 
            # 'codigosInterno': [codigointerno], 
            'versao': '1'
            # 'atributos': []
        })

        ncm_dict = next((x for x in attributes_json['listaNcm'] if str(x['codigoNcm']).replace('.', '') == ncm), None)

        for attr_dict in ncm_dict['listaAtributos']:

            attr_code = str(attr_dict['codigo']).strip()

            attr_dtls_dict = next((x for x in attributes_json['detalhesAtributos'] if str(x['codigo']).strip() == attr_code), None)

            attr_name = str(attr_dtls_dict['nomeApresentacao']).strip().lower()
            attr_value = None

            for attr in attribute_array:
                if attr['code'] == attr_code or attr['name'] in attr_name:
                    attr_value = attr['value']
                    break
        
            # results[-1]['atributos'].append({
            #     'atributo': attr_code, 
            #     'valor': attr_value
            # })

            if not attr_value:

                errors.append({
                    'seq': seq, 
                    'atributo': attr_code, 
                    'nome': attr_name, 
                    'valor': attr_value, 
                    'erro': 'Não foi possível associar um valor.'
                })

            else:
                if attr_dtls_dict['atributoCondicionante']:
                    for attr_cond_dict in attr_dtls_dict['condicionados']:

                        logic = attr_cond_dict['condicao']['operador']
                        condition_value = attr_cond_dict['condicao']['valor']

                        if eval(f'not "{attr_value}" {logic} "{condition_value}"'):
                            continue

                        attr_cond_code = str(attr_cond_dict['atributo']['codigo']).strip()
                        attr_cond_name = str(attr_cond_dict['atributo']['nomeApresentacao']).strip().lower()

                        attr_cond_value = None

                        for attr in attribute_array:
                            if attr['code'] == attr_cond_code or attr['name'] in attr_cond_name:
                                attr_cond_value = attr['value']
                                break
                    
                        # results[-1]['atributos'].append({
                        #     'atributo': attr_cond_code, 
                        #     'valor': attr_cond_value
                        # })

                        if not attr_cond_value:

                            errors.append({
                                'seq': seq, 
                                'atributo': attr_cond_code, 
                                'nome': attr_cond_name, 
                                'valor': attr_cond_value, 
                                'erro': 'Não foi possível associar um valor.'
                            })

    return results, errors


def makeProducts(products, attributes_json):
    
    results = []

    errors = []

    seq = 0

    for product in products:

        seq += 1

        ncm = str(product['ncm']).replace('.', '')
        descricao = str(product['descricao'])
        denominacao = str(product['denominacao'])
        raiz = (str(product['raiz']).replace('.', ''))[:8]
        # situacao = str(product['situacao']).upper()
        # modalidade = str(product['modalidade']).upper()
        codigointerno = str(product['codigoInterno']).strip()
        # codigo = str(product['códigoInterno']).strip()

        attribute_array = product['atributos']

        results.append({
            'seq': seq, 
            'descricao': descricao, 
            'denominacao': denominacao, 
            'cpfCnpjRaiz': raiz, 
            'situacao': 'ATIVADO', 
            'modalidade': 'IMPORTACAO', 
            'ncm': ncm, 
            'codigosInterno': [codigointerno], 
            'atributos': []
        })

        ncm_dict = next((x for x in attributes_json['listaNcm'] if str(x['codigoNcm']).replace('.', '') == ncm), None)

        for attr_dict in ncm_dict['listaAtributos']:

            attr_code = str(attr_dict['codigo']).strip()

            attr_dtls_dict = next((x for x in attributes_json['detalhesAtributos'] if str(x['codigo']).strip() == attr_code), None)

            attr_name = str(attr_dtls_dict['nomeApresentacao']).strip().lower()
            attr_value = None

            for attr in attribute_array:
                if attr['code'] == attr_code or attr['name'] in attr_name:
                    attr_value = attr['value']
                    break

            if str(attr_value).lower() == 'sim':
                attr_value = 'true'
            elif str(attr_value).lower().replace('ã', 'a') == 'nao':
                attr_value = 'false'
        
            results[-1]['atributos'].append({
                'atributo': attr_code, 
                'valor': attr_value
            })

            if not attr_value:

                errors.append({
                    'seq': seq, 
                    'atributo': attr_code, 
                    'ncm': ncm,
                    'condicional?': False, 
                    'nome': attr_name, 
                    'valor': attr_value, 
                    'erro': 'Não foi possível associar um valor.'
                })

            else:
                if attr_dtls_dict['atributoCondicionante']:
                    for attr_cond_dict in attr_dtls_dict['condicionados']:

                        logic = attr_cond_dict['condicao']['operador']
                        condition_value = attr_cond_dict['condicao']['valor']

                        if eval(f'not "{attr_value}" {logic} "{condition_value}"'):
                            continue

                        attr_cond_code = str(attr_cond_dict['atributo']['codigo']).strip()
                        attr_cond_name = str(attr_cond_dict['atributo']['nomeApresentacao']).strip().lower()

                        attr_cond_value = None

                        for attr in attribute_array:
                            if attr['code'] == attr_cond_code or attr['name'] in attr_cond_name:
                                attr_cond_value = attr['value']
                                break

                        if str(attr_cond_value).lower() == 'sim':
                            attr_cond_value = 'true'
                        elif str(attr_cond_value).lower().replace('ã', 'a') == 'nao':
                            attr_cond_value = 'false'
                    
                        results[-1]['atributos'].append({
                            'atributo': attr_cond_code, 
                            'valor': attr_cond_value
                        })

                        if not attr_cond_value:

                            errors.append({
                                'seq': seq, 
                                'atributo': attr_cond_code, 
                                'ncm': ncm,
                                'condicional?': True, 
                                'nome': attr_cond_name, 
                                'valor': attr_cond_value, 
                                'erro': 'Não foi possível associar um valor.'
                            })

    return results, errors

# products = excel_to_dict(df)

# data, errors = makeProducts(products, relation_json)

# Salvando erros, se houver
# if errors:
#     error_file = os.path.join(execution_folder, 'errors.xlsx')
#     pd.DataFrame(errors).to_excel(error_file, index=False)
#     print(f"Alguns erros ocorreram. Consulte o arquivo: {error_file}")

# # Salvando os dados de link, se houver
# data_file = os.path.join(execution_folder, 'link_data.xlsx')
# if data:
#     pd.DataFrame(data).to_excel(data_file, index=False)
#     print(f"Dados prontos para envio. Consulte o arquivo: {data_file}")